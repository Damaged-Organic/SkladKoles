import peewee
from openpyxl import load_workbook

db = peewee.MySQLDatabase(
    'cheersun_wheels', user='root', passwd='root', charset='utf8')


class SkladkolesItemsBrands(peewee.Model):
    brand = peewee.CharField()
    title = peewee.CharField()
    image = peewee.CharField()

    class Meta:
        database = db

    @classmethod
    def get_brands_data_source(cls, worksheet):
        items_list = []

        for row in worksheet.rows:
            for cell in row:
                formatted_brand = cell.value.lower().replace(' ', '_')

                items_list.append({
                    'brand': formatted_brand,
                    'title': cell.value,
                    'image': "{}.jpg".format(formatted_brand)
                })

        return items_list

    @classmethod
    def bulk_insert(cls, data_source):
        with db.atomic():
            cls.insert_many(data_source).execute()


class SkladkolesItemsBrandsRims(SkladkolesItemsBrands):
    class Meta:
        db_table = 'skladkoles_items_brands_rims'


class SkladkolesItemsBrandsTyres(SkladkolesItemsBrands):
    class Meta:
        db_table = 'skladkoles_items_brands_tyres'


# DROP/CREATE, then change collation to UTF-8, then INSERT

''' Rims '''

workbook_rims = load_workbook('diski_brand.xlsx', read_only=True)
worksheet_rims = workbook_rims.active

SkladkolesItemsBrandsRims.drop_table(fail_silently=True)
SkladkolesItemsBrandsRims.create_table(fail_silently=True)

db.execute_sql(
    "ALTER TABLE `skladkoles_items_brands_rims`"
    "CONVERT TO CHARACTER SET utf8 COLLATE utf8_general_ci;")

SkladkolesItemsBrandsRims.bulk_insert(
    SkladkolesItemsBrandsRims.get_brands_data_source(worksheet_rims))

''' Tyres '''

workbook_tyres = load_workbook('shini_brand.xlsx', read_only=True)
worksheet_tyres = workbook_tyres.active

SkladkolesItemsBrandsTyres.drop_table(fail_silently=True)
SkladkolesItemsBrandsTyres.create_table(fail_silently=True)

db.execute_sql(
    "ALTER TABLE `skladkoles_items_brands_tyres`"
    "CONVERT TO CHARACTER SET utf8 COLLATE utf8_general_ci;")

SkladkolesItemsBrandsTyres.bulk_insert(
    SkladkolesItemsBrandsTyres.get_brands_data_source(worksheet_tyres))
